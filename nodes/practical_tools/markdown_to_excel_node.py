import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import markdown

class MarkdownToExcelNode:
    """
    Markdown转Excel节点 - 将Markdown文件转换为Excel文件，支持表格解析
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "md_file_path": ("STRING", {"default": "", "multiline": False}),
                "output_dir": ("STRING", {"default": "", "multiline": False}),
                "output_filename": ("STRING", {"default": "", "multiline": False}),
                "sheet_name": ("STRING", {"default": "Sheet1", "multiline": False}),
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("conversion_info",)
    
    FUNCTION = "convert_md_to_excel"
    
    CATEGORY = "XnanTool/实用工具"

    def convert_md_to_excel(self, md_file_path, output_dir, output_filename, sheet_name="Sheet1"):
        """
        将Markdown文件转换为Excel文件，支持表格解析
        
        Args:
            md_file_path: Markdown文件路径
            output_dir: 输出目录路径
            output_filename: 输出文件名
            sheet_name: 工作表名称，默认为"Sheet1"
            
        Returns:
            tuple: (转换信息,)
        """
        try:
            # 检查输入文件是否存在
            if not os.path.exists(md_file_path):
                raise FileNotFoundError(f"Markdown文件不存在: {md_file_path}")
            
            # 获取输入文件的基本信息
            input_dir, input_filename = os.path.split(md_file_path)
            input_basename, _ = os.path.splitext(input_filename)
            
            # 确定输出目录
            if not output_dir:
                output_dir = input_dir  # 如果没有指定输出目录，则使用输入文件所在目录
            
            # 确定输出文件名
            if not output_filename:
                output_filename = input_basename + ".xlsx"  # 如果没有指定输出文件名，则使用输入文件名
            elif not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'  # 确保文件扩展名为.xlsx
            
            # 构建完整的输出文件路径
            output_excel_path = os.path.join(output_dir, output_filename)
            
            # 读取Markdown文件内容
            with open(md_file_path, 'r', encoding='utf-8') as file:
                md_content = file.read()
            
            # 创建一个工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name  # 使用指定的工作表名称
            
            # 按行处理Markdown内容
            lines = md_content.split('\n')
            row_idx = 1
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                if not line:
                    # 空行
                    row_idx += 1
                    i += 1
                    continue
                
                # 检查是否是表格行
                if line.startswith('|') and '|' in line:
                    # 这是一个表格，解析整个表格
                    table_rows = []
                    while i < len(lines) and lines[i].strip().startswith('|'):
                        # 解析表格行
                        cells = [cell.strip() for cell in lines[i].strip().split('|')]
                        # 移除首尾空元素
                        cells = [cell for cell in cells if cell != '']
                        table_rows.append(cells)
                        i += 1
                    
                    # 将表格写入工作表
                    for table_row in table_rows:
                        ws.append(table_row)
                        row_idx += 1
                elif line.startswith('# '):
                    # 一级标题
                    ws.cell(row=row_idx, column=1, value=line[2:])
                    row_idx += 1
                    i += 1
                elif line.startswith('## '):
                    # 二级标题
                    ws.cell(row=row_idx, column=1, value=line[3:])
                    row_idx += 1
                    i += 1
                elif line.startswith('### '):
                    # 三级标题
                    ws.cell(row=row_idx, column=1, value=line[4:])
                    row_idx += 1
                    i += 1
                elif line.startswith('- ') or line.startswith('* '):
                    # 列表项
                    ws.cell(row=row_idx, column=1, value=line)
                    row_idx += 1
                    i += 1
                else:
                    # 普通段落
                    ws.cell(row=row_idx, column=1, value=line)
                    row_idx += 1
                    i += 1
            
            # 保存Excel文件
            wb.save(output_excel_path)
            
            conversion_info = f"成功将 {md_file_path} 转换为 {output_excel_path} (工作表: {sheet_name})"
            return (conversion_info,)
            
        except Exception as e:
            error_msg = f"转换失败: {str(e)}"
            return (error_msg,)


# 注册节点
NODE_CLASS_MAPPINGS = {
    "MarkdownToExcelNode": MarkdownToExcelNode
}

NODE_DISPLAY_NAME_MAPPINGS = {
    "MarkdownToExcelNode": "MD转Excel"
}

__all__ = ['NODE_CLASS_MAPPINGS', 'NODE_DISPLAY_NAME_MAPPINGS']